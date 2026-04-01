from pathlib import Path
from urllib.parse import urlparse
from datetime import date
from typing_extensions import TypedDict
import gzip
import os
import re
import requests
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook
from trafilatura import fetch_url, extract
from langgraph.graph import StateGraph, START, END


# ---------------------------
# Project folders
# ---------------------------
DOWNLOAD_DIR = Path("download_xml")
EXCEL_DIR = Path("extracted_urls")
MARKDOWN_DIR = Path("scraped_markdown")
DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0 Safari/537.36"
    )
}


class PipelineState(TypedDict, total=False):
    # Input values
    base_url: str
    cutoff_date: str

    # File paths created by nodes
    xml_file_path: str
    excel_file_path: str


# ---------------------------
# Small helper functions
# ---------------------------
def get_site_name(url: str) -> str:
    """Extract website name from a base URL.

    Example:
    https://www.motortrend.com/ -> motortrend
    """
    netloc = urlparse(url).netloc.lower()
    if netloc.startswith("www."):
        netloc = netloc[4:]
    return netloc.split(".")[0]


def make_safe_slug(value: str, fallback: str) -> str:
    """Convert any text into a file-safe slug."""
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-")
    return (slug or fallback)[:180]


def build_sitemap_url(base_url: str) -> str:
    """Always download the top sitemap from the website root."""
    return base_url.rstrip("/") + "/sitemap.xml"


def to_date(text: str) -> date | None:
    """Convert sitemap lastmod text into a Python date.

    Most sitemap lastmod values are ISO-like strings, so taking the first
    10 characters is enough for values like:
    2026-03-25
    2026-03-25T10:30:00Z
    """
    text = str(text or "").strip()
    if not text:
        return None
    return date.fromisoformat(text[:10])


def date_in_url(url: str) -> date | None:
    """Fallback date extractor from child sitemap URLs."""
    match = DATE_RE.search(url)
    if not match:
        return None
    return date.fromisoformat(match.group(1))


def xml_root_from_bytes(content: bytes, source: str = "") -> ET.Element:
    """Read sitemap XML bytes, including .gz child sitemaps."""
    is_gzip = source.lower().endswith(".gz") or content[:2] == b"\x1f\x8b"
    if is_gzip:
        content = gzip.decompress(content)

    text = content.decode("utf-8-sig", errors="replace").strip()
    return ET.fromstring(text)


def read_xml_root_from_file(xml_file_path: str) -> ET.Element:
    """Read the already-downloaded top sitemap file."""
    content = Path(xml_file_path).read_bytes()
    return xml_root_from_bytes(content, xml_file_path)


def get_namespace(root: ET.Element) -> dict[str, str]:
    """Handle namespace-aware sitemap parsing."""
    if root.tag.startswith("{"):
        return {"sm": root.tag.split("}", 1)[0][1:]}
    return {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}


def get_root_tag_name(root: ET.Element) -> str:
    """Return root tag without namespace.

    Examples:
    {namespace}urlset -> urlset
    {namespace}sitemapindex -> sitemapindex
    """
    if "}" in root.tag:
        return root.tag.split("}", 1)[1]
    return root.tag


def get_child_text(node: ET.Element, tag: str, namespace: dict[str, str]) -> str:
    """Read text from a child XML tag safely."""
    child = node.find(f"sm:{tag}", namespace)
    return child.text.strip() if child is not None and child.text else ""


def keep_url(lastmod: str, cutoff: date) -> bool:
    """Keep only final URLs whose lastmod is on/after cutoff date."""
    parsed = to_date(lastmod)
    return parsed is not None and parsed >= cutoff


def expand_child(child_url: str, parent_lastmod: str, cutoff: date) -> bool:
    """Decide whether to open a child sitemap.

    The child sitemap is opened only if:
    - parent lastmod exists and is recent enough, or
    - the child sitemap URL itself contains a recent enough date.
    """
    parsed_parent = to_date(parent_lastmod)
    parsed_from_url = date_in_url(child_url)
    effective_date = parsed_parent or parsed_from_url
    return effective_date is not None and effective_date >= cutoff


def get_article_slug(url: str) -> str:
    """Build an article-like file name from the URL path."""
    parsed = urlparse(url)
    path_parts = [part for part in parsed.path.split("/") if part]

    if path_parts:
        candidate = path_parts[-1]
    else:
        candidate = parsed.netloc

    return make_safe_slug(candidate, "article")


def to_relative_path(file_path: Path) -> str:
    """Store relative paths in Excel instead of long absolute paths."""
    return os.path.relpath(file_path.resolve(), start=Path.cwd().resolve()).replace("\\", "/")


# ---------------------------
# Node 1: Download top sitemap XML
# ---------------------------
def download_xml(state: PipelineState) -> dict:
    """Download top sitemap.xml and save it locally."""
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    base_url = state["base_url"]
    sitemap_url = build_sitemap_url(base_url)

    site_name = get_site_name(base_url)
    file_path = DOWNLOAD_DIR / f"{site_name}.xml"

    response = requests.get(sitemap_url, headers=HEADERS, timeout=30)
    response.raise_for_status()

    file_path.write_bytes(response.content)

    return {"xml_file_path": str(file_path.resolve())}


# ---------------------------
# Recursive sitemap walker
# ---------------------------
def walk_sitemap(
    root: ET.Element,
    session: requests.Session,
    cutoff: date,
    final_rows: list[dict],
    seen_sitemaps: set[str],
    seen_links: set[str],
) -> None:
    """Recursively walk sitemaps until final article URLs are reached.

    - If root is urlset: collect final article URLs.
    - If root is sitemapindex: open child sitemap files and recurse.
    """
    namespace = get_namespace(root)
    root_tag = get_root_tag_name(root)

    if root_tag == "urlset":
        for node in root.findall("./sm:url", namespace):
            link = get_child_text(node, "loc", namespace)
            lastmod = get_child_text(node, "lastmod", namespace)

            if link and keep_url(lastmod, cutoff) and link not in seen_links:
                seen_links.add(link)
                final_rows.append(
                    {
                        "link": link,
                        "lastmod": lastmod,
                    }
                )
        return

    if root_tag == "sitemapindex":
        for node in root.findall("./sm:sitemap", namespace):
            child_url = get_child_text(node, "loc", namespace)
            parent_lastmod = get_child_text(node, "lastmod", namespace)

            if not child_url:
                continue

            if child_url in seen_sitemaps:
                continue

            if not expand_child(child_url, parent_lastmod, cutoff):
                continue

            seen_sitemaps.add(child_url)

            response = session.get(child_url, headers=HEADERS, timeout=30)
            response.raise_for_status()

            child_root = xml_root_from_bytes(response.content, child_url)
            walk_sitemap(child_root, session, cutoff, final_rows, seen_sitemaps, seen_links)


# ---------------------------
# Node 2: Extract URLs and save them to Excel
# ---------------------------
def extract_urls_to_excel(state: PipelineState) -> dict:
    """Recursively extract final article URLs and save them to an Excel file.

    The Excel file starts with these columns:
    - link
    - lastmod
    - markdown_path
    - markdown_status
    """
    cutoff = date.fromisoformat(state["cutoff_date"])
    top_root = read_xml_root_from_file(state["xml_file_path"])
    site_name = get_site_name(state["base_url"])

    final_rows: list[dict] = []
    seen_sitemaps: set[str] = set()
    seen_links: set[str] = set()

    with requests.Session() as session:
        session.headers.update(HEADERS)
        walk_sitemap(
            root=top_root,
            session=session,
            cutoff=cutoff,
            final_rows=final_rows,
            seen_sitemaps=seen_sitemaps,
            seen_links=seen_links,
        )

    # Create site-specific Excel folder and file.
    site_excel_dir = EXCEL_DIR / site_name
    site_excel_dir.mkdir(parents=True, exist_ok=True)
    excel_path = site_excel_dir / f"{site_name}_urls.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = site_name[:31]

    # Write header row.
    worksheet.append(["link", "lastmod", "markdown_path", "markdown_status"])

    # Write extracted sitemap rows.
    for row in final_rows:
        worksheet.append([row["link"], row["lastmod"], "", ""])

    workbook.save(excel_path)
    workbook.close()

    return {"excel_file_path": str(excel_path.resolve())}


# ---------------------------
# Node 3: Download article markdown and update Excel
# ---------------------------
def download_markdown_from_excel(state: PipelineState) -> dict:
    """Read URLs from Excel, extract markdown with Trafilatura, save files,
    and update the Excel workbook with markdown path + markdown status.

    Status values used:
    - success : markdown extracted and saved
    - failed  : fetch/extract raised an error
    - empty   : extraction completed but markdown was empty
    """
    excel_path = Path(state["excel_file_path"])
    site_name = get_site_name(state["base_url"])

    # Create site-specific markdown folder.
    site_markdown_dir = MARKDOWN_DIR / site_name
    site_markdown_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    # Map headers to column indexes.
    header_map = {}
    for col in range(1, worksheet.max_column + 1):
        header_value = str(worksheet.cell(row=1, column=col).value or "").strip().lower()
        if header_value:
            header_map[header_value] = col

    link_col = header_map["link"]
    markdown_path_col = header_map["markdown_path"]
    markdown_status_col = header_map["markdown_status"]

    for row_number in range(2, worksheet.max_row + 1):
        link = str(worksheet.cell(row=row_number, column=link_col).value or "").strip()

        if not link:
            worksheet.cell(row=row_number, column=markdown_status_col, value="empty")
            workbook.save(excel_path)
            continue

        try:
            downloaded = fetch_url(link)

            if not downloaded:
                worksheet.cell(row=row_number, column=markdown_path_col, value="")
                worksheet.cell(row=row_number, column=markdown_status_col, value="failed")
                workbook.save(excel_path)
                continue

            markdown = extract(
                downloaded,
                url=link,
                output_format="markdown",
                with_metadata=True,
                include_tables=True,
                include_images=True,
                include_links=True,
                favor_recall=True,
                deduplicate=True,
            )

            if not markdown or not markdown.strip():
                worksheet.cell(row=row_number, column=markdown_path_col, value="")
                worksheet.cell(row=row_number, column=markdown_status_col, value="empty")
                workbook.save(excel_path)
                continue

            # Use article name from the URL path as file name.
            article_slug = get_article_slug(link)
            markdown_file_path = site_markdown_dir / f"{article_slug}.md"

            # If the same slug appears multiple times, add row number to keep files unique.
            if markdown_file_path.exists():
                markdown_file_path = site_markdown_dir / f"{article_slug}_{row_number}.md"

            markdown_file_path.write_text(markdown, encoding="utf-8")

            worksheet.cell(
                row=row_number,
                column=markdown_path_col,
                value=to_relative_path(markdown_file_path),
            )
            worksheet.cell(row=row_number, column=markdown_status_col, value="success")

        except Exception:
            worksheet.cell(row=row_number, column=markdown_path_col, value="")
            worksheet.cell(row=row_number, column=markdown_status_col, value="failed")

        # Save after every row so partial progress is never lost.
        workbook.save(excel_path)

    workbook.close()
    return {"excel_file_path": str(excel_path.resolve())}


# ---------------------------
# LangGraph pipeline
# ---------------------------
builder = StateGraph(PipelineState)

# Add nodes.
builder.add_node("download_xml", download_xml)
builder.add_node("extract_urls_to_excel", extract_urls_to_excel)
builder.add_node("download_markdown_from_excel", download_markdown_from_excel)

# Run everything sequentially.
builder.add_edge(START, "download_xml")
builder.add_edge("download_xml", "extract_urls_to_excel")
builder.add_edge("extract_urls_to_excel", "download_markdown_from_excel")
builder.add_edge("download_markdown_from_excel", END)

graph = builder.compile()


if __name__ == "__main__":
    result = graph.invoke(
        {
            "base_url": "https://www.automotiveworld.com/",
            "cutoff_date": "2026-03-25",
        }
    )

    print("Downloaded XML:", result["xml_file_path"])
    print("Excel file:", result["excel_file_path"])
