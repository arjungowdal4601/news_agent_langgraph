from pathlib import Path

from openpyxl import Workbook, load_workbook as openpyxl_load_workbook


def create_workbook() -> Workbook:
    """Create a fresh workbook for extracted sitemap rows."""
    return Workbook()


def write_headers(worksheet) -> None:
    """Write the Excel header row used by the markdown step."""
    worksheet.append(["link", "lastmod", "markdown_path", "markdown_status"])


def append_extracted_rows(worksheet, rows: list[dict]) -> None:
    """Append the extracted sitemap rows to the worksheet."""
    for row in rows:
        worksheet.append([row["link"], row["lastmod"], "", ""])


def load_workbook(excel_path: Path) -> Workbook:
    """Open an existing workbook so markdown results can be written back."""
    return openpyxl_load_workbook(excel_path)


def map_header_columns(worksheet) -> dict[str, int]:
    """Map lowercase header names to their Excel column indexes."""
    header_map: dict[str, int] = {}

    for col in range(1, worksheet.max_column + 1):
        header_value = str(worksheet.cell(row=1, column=col).value or "").strip().lower()
        if header_value:
            header_map[header_value] = col

    return header_map


def ensure_columns(worksheet, required_headers: list[str]) -> dict[str, int]:
    """Add missing headers to the worksheet and return the updated header map."""
    header_map = map_header_columns(worksheet)

    for header in required_headers:
        if header not in header_map:
            worksheet.cell(row=1, column=worksheet.max_column + 1, value=header)
            header_map = map_header_columns(worksheet)

    return header_map


def update_markdown_path(worksheet, row_number: int, column: int, value: str) -> None:
    """Write the markdown file path for a row."""
    worksheet.cell(row=row_number, column=column, value=value)


def update_markdown_status(worksheet, row_number: int, column: int, value: str) -> None:
    """Write the markdown extraction status for a row."""
    worksheet.cell(row=row_number, column=column, value=value)


def update_similarity_score(worksheet, row_number: int, column: int, value) -> None:
    """Write the semantic similarity score for a row."""
    worksheet.cell(row=row_number, column=column, value=value)


def update_semantic_status(worksheet, row_number: int, column: int, value: str) -> None:
    """Write the semantic routing status for a row."""
    worksheet.cell(row=row_number, column=column, value=value)


def update_semantic_reason(worksheet, row_number: int, column: int, value: str) -> None:
    """Write the semantic routing reason for a row."""
    worksheet.cell(row=row_number, column=column, value=value)


def update_final_markdown_path(worksheet, row_number: int, column: int, value: str) -> None:
    """Write the final newsletter markdown path for a row."""
    worksheet.cell(row=row_number, column=column, value=value)


def update_final_markdown_status(worksheet, row_number: int, column: int, value: str) -> None:
    """Write the final newsletter composition status for a row."""
    worksheet.cell(row=row_number, column=column, value=value)


def save_workbook(workbook: Workbook, excel_path: Path) -> None:
    """Persist workbook updates after each row."""
    workbook.save(excel_path)
